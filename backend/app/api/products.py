import io
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.database import get_db
from app.dependencies import require_permission
from app.models.product import Product
from app.models.user import User
from app.schemas.product import (
    CreateProductRequest,
    ProductListResponse,
    ProductResponse,
    UpdateProductRequest,
)

router = APIRouter(prefix="/admin/products", tags=["products"])

# Excel column header → DB field mapping
EXCEL_COLUMN_MAP = {
    "Ürün Kodu": "urun_kodu",
    "Marka": "marka",
    "Koleksiyon": "koleksiyon",
    "Ürün Tanımı": "urun_tanimi",
    "Ürün Tipi": "urun_tipi",
    "Ebat (cm)": "ebat_cm",
    "Hacim (cc)": "hacim_cc",
    "Servis Tipi": "servis_tipi",
    "Yemek Öneri": "yemek_onerileri",
    "Menü Ana Başlık": "menu_ana_baslik",
    "Fiyat Segmenti": "fiyat_segmenti",
    "Segment": "segment",
    "Kullanım Alanı": "kullanim_alani",
    "Ana Renk": "ana_renk",
    "Kenar Çıtlama Dayanımı": "kenar_citlama_dayanimi",
    "Stil": "stil",
    "Mutfak Uyumu": "mutfak_uyumu",
    "Dekor Kodu": "dekor",
}


@router.get("/template")
async def download_template(
    user: Annotated[User, Depends(require_permission("documents.view"))],
):
    """Download an empty Excel template for product import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    headers = list(EXCEL_COLUMN_MAP.keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=urun-sablonu.xlsx"},
    )


@router.post("/import")
async def import_products(
    user: Annotated[User, Depends(require_permission("documents.upload"))],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
):
    """Import products from an Excel file. Upsert by urun_kodu."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Sadece .xlsx dosyaları kabul edilir")

    from openpyxl import load_workbook

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Excel dosyası okunamadı")

    # Read header row and build column index mapping
    rows = ws.iter_rows(min_row=1)
    header_row = next(rows, None)
    if not header_row:
        raise HTTPException(400, "Excel dosyası boş")

    col_map: dict[int, str] = {}  # col_index → db_field
    for idx, cell in enumerate(header_row):
        val = str(cell.value).strip() if cell.value else ""
        if val in EXCEL_COLUMN_MAP:
            col_map[idx] = EXCEL_COLUMN_MAP[val]

    if "urun_kodu" not in col_map.values():
        raise HTTPException(400, "'Ürün Kodu' kolonu bulunamadı")

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        try:
            row_data: dict[str, str | None] = {}
            for col_idx, db_field in col_map.items():
                cell_val = row[col_idx].value if col_idx < len(row) else None
                if cell_val is None:
                    continue
                val_str = str(cell_val).strip()
                if val_str == "" or val_str == "-":
                    continue
                # Special handling for hacim_cc (integer)
                if db_field == "hacim_cc":
                    try:
                        row_data[db_field] = int(float(val_str))
                    except (ValueError, TypeError):
                        continue
                else:
                    row_data[db_field] = val_str

            urun_kodu = row_data.get("urun_kodu")
            if not urun_kodu:
                skipped += 1
                continue

            # Check if product exists
            result = await db.execute(
                select(Product).where(Product.urun_kodu == str(urun_kodu))
            )
            existing = result.scalar_one_or_none()

            if existing:
                # Update only non-empty fields (skip urun_kodu itself)
                update_fields = {k: v for k, v in row_data.items() if k != "urun_kodu"}
                if update_fields:
                    for key, value in update_fields.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    skipped += 1
            else:
                product = Product(**row_data)
                db.add(product)
                created += 1

        except Exception as e:
            errors.append(f"Satır {row_num}: {str(e)}")
            if len(errors) > 50:
                errors.append("...daha fazla hata var")
                break

    await db.commit()
    wb.close()

    return {
        "status": "ok",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


@router.get("", response_model=ProductListResponse)
async def list_products(
    user: Annotated[User, Depends(require_permission("documents.view"))],
    db: Annotated[AsyncSession, Depends(get_db)],
    search: str | None = Query(None),
    marka: str | None = Query(None),
    koleksiyon: str | None = Query(None),
    urun_tipi: str | None = Query(None),
    aktif: bool | None = Query(None),
    limit: int = Query(20, le=100),
    offset: int = Query(0, ge=0),
):
    """List products with pagination, search, and filters."""
    query = select(Product)
    count_query = select(func.count(Product.id))

    if search:
        search_filter = or_(
            Product.urun_kodu.ilike(f"%{search}%"),
            Product.urun_tanimi.ilike(f"%{search}%"),
            Product.marka.ilike(f"%{search}%"),
            Product.koleksiyon.ilike(f"%{search}%"),
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)
    if marka:
        query = query.where(Product.marka == marka)
        count_query = count_query.where(Product.marka == marka)
    if koleksiyon:
        query = query.where(Product.koleksiyon == koleksiyon)
        count_query = count_query.where(Product.koleksiyon == koleksiyon)
    if urun_tipi:
        query = query.where(Product.urun_tipi == urun_tipi)
        count_query = count_query.where(Product.urun_tipi == urun_tipi)
    if aktif is not None:
        query = query.where(Product.aktif == aktif)
        count_query = count_query.where(Product.aktif == aktif)

    total = (await db.execute(count_query)).scalar() or 0
    result = await db.execute(
        query.order_by(Product.id.desc()).limit(limit).offset(offset)
    )
    products = result.scalars().all()

    return ProductListResponse(
        products=[ProductResponse.model_validate(p) for p in products],
        total=total,
    )


@router.get("/filters")
async def get_product_filters(
    user: Annotated[User, Depends(require_permission("documents.view"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get distinct values for filter dropdowns."""
    markalar = (await db.execute(
        select(Product.marka).where(Product.marka.isnot(None)).distinct().order_by(Product.marka)
    )).scalars().all()
    koleksiyonlar = (await db.execute(
        select(Product.koleksiyon).where(Product.koleksiyon.isnot(None)).distinct().order_by(Product.koleksiyon)
    )).scalars().all()
    urun_tipleri = (await db.execute(
        select(Product.urun_tipi).where(Product.urun_tipi.isnot(None)).distinct().order_by(Product.urun_tipi)
    )).scalars().all()
    return {
        "markalar": markalar,
        "koleksiyonlar": koleksiyonlar,
        "urun_tipleri": urun_tipleri,
    }


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    user: Annotated[User, Depends(require_permission("documents.view"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get single product detail."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Ürün bulunamadı")
    return ProductResponse.model_validate(product)


@router.post("", response_model=ProductResponse)
async def create_product(
    body: CreateProductRequest,
    user: Annotated[User, Depends(require_permission("documents.upload"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Create a new product."""
    # Check duplicate urun_kodu
    existing = await db.execute(
        select(Product).where(Product.urun_kodu == body.urun_kodu)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(409, "Bu ürün kodu zaten mevcut")

    product = Product(**body.model_dump())
    db.add(product)
    await db.flush()
    await db.commit()
    await db.refresh(product)

    return ProductResponse.model_validate(product)


@router.put("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    body: UpdateProductRequest,
    user: Annotated[User, Depends(require_permission("documents.upload"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Update a product."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Ürün bulunamadı")

    # Check duplicate urun_kodu if changing
    if body.urun_kodu is not None and body.urun_kodu != product.urun_kodu:
        existing = await db.execute(
            select(Product).where(
                Product.urun_kodu == body.urun_kodu, Product.id != product_id
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(409, "Bu ürün kodu zaten mevcut")

    update_data = body.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(product, key, value)

    await db.flush()
    await db.commit()
    await db.refresh(product)

    return ProductResponse.model_validate(product)


@router.delete("/{product_id}")
async def delete_product(
    product_id: int,
    user: Annotated[User, Depends(require_permission("documents.delete"))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Delete a product."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Ürün bulunamadı")

    await db.delete(product)
    await db.commit()

    return {"status": "ok", "message": "Ürün silindi"}
